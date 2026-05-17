import argparse
import asyncio
import threading
import time
from datetime import datetime
from pathlib import Path

from bleak import BleakClient, BleakScanner
from openpyxl import Workbook, load_workbook
import serial

FRAME_HEADER = b"\xD5\x5D"
PRESSURE_POINTS = 24
PRESSURE_BYTES = PRESSURE_POINTS * 2
PAYLOAD_LEN = 57
DEFAULT_OUTPUT_XLSX = "pressure_log.xlsx"
# 三次拟合函数
# weight = a*x^3 + b*x^2 + c*x + d，其中 x 为 24 点总和
WEIGHT_FIT_A = -3.7086554e-11
WEIGHT_FIT_B = 2.9099358e-06
WEIGHT_FIT_C = -0.070669315
WEIGHT_FIT_D = 553.75605


def normalize_uuid(value: str) -> str:
    uuid = value.strip().lower()
    if len(uuid) == 4:
        return f"0000{uuid}-0000-1000-8000-00805f9b34fb"
    return uuid


def format_bytes(data: bytearray) -> str:
    return " ".join(f"{byte:02X}" for byte in data)


def parse_pressure_packet(data: bytearray) -> tuple[list[int], int]:
    raw = bytes(data)
    min_packet_len = 2 + 1 + PAYLOAD_LEN + 2
    if len(raw) < min_packet_len:
        raise ValueError(f"数据长度不足: {len(raw)} < {min_packet_len}")

    if raw[0:2] != FRAME_HEADER:
        raise ValueError(f"帧头错误: {raw[0:2].hex()}")

    payload_len = raw[2]
    if payload_len != PAYLOAD_LEN:
        raise ValueError(f"数据长度字段异常: {payload_len} != {PAYLOAD_LEN}")

    payload_end = 3 + payload_len
    payload = raw[3:payload_end]

    sensor_id = int.from_bytes(payload[0:4], "little")
    pressure_data = payload[4: 4 + PRESSURE_BYTES]
    points = [
        int.from_bytes(pressure_data[i: i + 2], "little")
        for i in range(0, PRESSURE_BYTES, 2)
    ]
    total = sum(points)

    print(f"传感器编号: 0x{sensor_id:08X}")
    return points, total


def estimate_weight(total: int) -> float:
    return (
        WEIGHT_FIT_A * (total ** 3)
        + WEIGHT_FIT_B * (total ** 2)
        + WEIGHT_FIT_C * total
        + WEIGHT_FIT_D
    )


def parse_scale_weight(data: bytes) -> float | None:
    # 例如: b"\x2D\x30\x2E\x31\x37\x2C\x6B\x67\x2C\x0D\x0A" -> "-0.17,kg,"
    text = data.decode("ascii", errors="ignore").strip()
    if not text:
        return None
    value_text = text.split(",", 1)[0].strip()
    try:
        return float(value_text)
    except ValueError:
        return None


class SerialScaleReader:
    def __init__(self, port: str, baudrate: int = 9600) -> None:
        self.port = port
        self.baudrate = baudrate
        self._stop_event = threading.Event()
        self._thread: threading.Thread | None = None
        self._latest_weight: float | None = None
        self._latest_hex = ""
        self._lock = threading.Lock()

    def start(self) -> None:
        self._thread = threading.Thread(target=self._read_loop, daemon=True)
        self._thread.start()

    def stop(self) -> None:
        self._stop_event.set()
        if self._thread is not None:
            self._thread.join(timeout=2.0)

    def get_latest_weight(self) -> float | None:
        with self._lock:
            return self._latest_weight

    def get_latest_hex(self) -> str:
        with self._lock:
            return self._latest_hex

    def _read_loop(self) -> None:
        try:
            with serial.Serial(self.port, self.baudrate, timeout=1) as ser:
                print(f"串口体重计已连接: {self.port} @ {self.baudrate}")
                while not self._stop_event.is_set():
                    raw = ser.readline()
                    if not raw:
                        continue

                    weight = parse_scale_weight(raw)
                    raw_hex = " ".join(f"{b:02X}" for b in raw)
                    with self._lock:
                        self._latest_hex = raw_hex
                        if weight is not None:
                            self._latest_weight = weight

                    if weight is not None:
                        print(f"[Scale] 原始HEX: {raw_hex} | 解析重量: {weight:.3f} kg")
                    else:
                        print(f"[Scale] 原始HEX: {raw_hex} | 解析失败")
        except Exception as exc:
            print(f"串口读取失败({self.port}): {exc}")


def append_pressure_to_excel(
    timestamp: str,
    points: list[int],
    total: int,
    weight: float,
    scale_weight: float | None,
    output_path: str = DEFAULT_OUTPUT_XLSX,
) -> None:
    file_path = Path(output_path)
    headers = ["时间"] + [f"点{i}" for i in range(1, PRESSURE_POINTS + 1)] + ["总和", "换算重量", "串口重量", "误差(换算-串口)"]
    diff = weight - scale_weight if scale_weight is not None else None
    row = [timestamp] + points + [total, weight, scale_weight, diff]

    if file_path.exists():
        workbook = load_workbook(file_path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "pressure_log"
        worksheet.append(headers)

    worksheet.append(row)
    workbook.save(file_path)


async def scan_ble_devices(timeout: float, name_keyword: str) -> set[str]:
    print(f"开始扫描附近 BLE 设备，时长 {timeout} 秒...\n")
    devices = await BleakScanner.discover(timeout=timeout, return_adv=True)

    if not devices:
        print("未发现任何 BLE 设备。")
        return set()

    keyword_lower = name_keyword.lower()
    matched_addresses: set[str] = set()
    print(f"扫描完成，共发现 {len(devices)} 个设备。")
    print(f"以下是名称包含关键字 '{name_keyword}' 的设备：\n")

    for index, (address, (device, adv_data)) in enumerate(devices.items(), start=1):
        name = device.name or adv_data.local_name or "未知设备"
        rssi = adv_data.rssi
        if keyword_lower in name.lower():
            matched_addresses.add(address.upper())
            print(f"{index}. 名称: {name}")
            print(f"   地址: {address}")
            print(f"   RSSI: {rssi} dBm")
            print("-" * 40)

    if not matched_addresses:
        print("未找到名称包含该关键字的设备。")

    return matched_addresses


async def connect_by_mac(
    mac_address: str,
    read_uuid: str,
    notify_uuid: str,
    notify_seconds: float,
    output_xlsx: str,
    scale_reader: SerialScaleReader | None,
) -> None:
    print(f"\n尝试连接设备: {mac_address}")
    async with BleakClient(mac_address) as client:
        if client.is_connected:
            print("连接成功！")
            service_count = 0
            if hasattr(client, "get_services"):
                services = await client.get_services()
                service_count = len(list(services))
            else:
                services = getattr(client, "services", None)
                if services is not None:
                    service_count = len(list(services))
            print(f"已发现服务数量: {service_count}")

            try:
                read_data = await client.read_gatt_char(read_uuid)
                print(f"读取 {read_uuid} 成功: {format_bytes(read_data)}")
            except Exception as exc:
                print(f"读取 {read_uuid} 失败: {exc}")

            def notification_handler(sender: int, data: bytearray) -> None:
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"[{ts}] [Notify] {notify_uuid} <- {format_bytes(data)} (handle={sender})")
                try:
                    points, total = parse_pressure_packet(data)
                    weight = estimate_weight(total)
                    scale_weight = scale_reader.get_latest_weight() if scale_reader is not None else None
                    print(f"[{ts}] 24点压力值: {points}")
                    print(f"[{ts}] 压力值总和: {total}")
                    print(f"[{ts}] 换算重量: {weight:.3f}")
                    if scale_weight is not None:
                        print(f"[{ts}] 串口重量: {scale_weight:.3f} kg")
                        print(f"[{ts}] 重量误差(换算-串口): {weight - scale_weight:.3f}")
                    else:
                        print(f"[{ts}] 串口重量: 暂无有效数据")
                    append_pressure_to_excel(ts, points, total, weight, scale_weight, output_xlsx)
                    print(f"[{ts}] 已写入 Excel: {output_xlsx}")
                except Exception as exc:
                    print(f"[{ts}] 解析失败: {exc}")

            try:
                await client.start_notify(notify_uuid, notification_handler)
                print(f"已开启 {notify_uuid} notify，监听 {notify_seconds} 秒...")
                await asyncio.sleep(notify_seconds)
                await client.stop_notify(notify_uuid)
                print(f"已关闭 {notify_uuid} notify")
            except Exception as exc:
                print(f"开启或接收 {notify_uuid} notify 失败: {exc}")
        else:
            print("连接失败。")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="扫描并连接 BLE 蓝牙设备")
    parser.add_argument(
        "--timeout",
        type=float,
        default=6.0,
        help="扫描时长（秒），默认 8 秒",
    )
    parser.add_argument(
        "--keyword",
        type=str,
        default="WM",
        help="设备名匹配关键字，默认 WM",
    )
    parser.add_argument(
        "--read-uuid",
        type=str,
        default="FFE1",
        help="连接后读取的特征 UUID，默认 FFE1",
    )
    parser.add_argument(
        "--notify-uuid",
        type=str,
        default="FFE4",
        help="连接后开启 notify 的特征 UUID，默认 FFE4",
    )
    parser.add_argument(
        "--notify-seconds",
        type=float,
        default=60.0,
        help="notify 监听时长（秒），默认 30 秒",
    )
    parser.add_argument(
        "--output-xlsx",
        type=str,
        default=DEFAULT_OUTPUT_XLSX,
        help="数据保存的 Excel 文件路径，默认 pressure_log.xlsx",
    )
    parser.add_argument(
        "--scale-port",
        type=str,
        default="",
        help="体重计串口号（例如 COM3），不填则不启用串口读取",
    )
    parser.add_argument(
        "--scale-baudrate",
        type=int,
        default=9600,
        help="体重计串口波特率，默认 9600",
    )
    return parser.parse_args()


async def run() -> None:
    args = parse_args()
    scale_reader: SerialScaleReader | None = None
    if args.scale_port:
        scale_reader = SerialScaleReader(args.scale_port, args.scale_baudrate)
        scale_reader.start()
        # 给串口线程一点启动时间
        time.sleep(0.2)

    matched_addresses = await scan_ble_devices(args.timeout, args.keyword)
    if not matched_addresses:
        if scale_reader is not None:
            scale_reader.stop()
        return

    mac_address = input("\n请输入要连接的设备 MAC 地址: ").strip().upper()
    if not mac_address:
        print("未输入 MAC 地址，已退出。")
        if scale_reader is not None:
            scale_reader.stop()
        return

    read_uuid = normalize_uuid(args.read_uuid)
    notify_uuid = normalize_uuid(args.notify_uuid)
    try:
        await connect_by_mac(
            mac_address,
            read_uuid,
            notify_uuid,
            args.notify_seconds,
            args.output_xlsx,
            scale_reader,
        )
    finally:
        if scale_reader is not None:
            scale_reader.stop()


def main() -> None:
    try:
        asyncio.run(run())
    except ImportError:
        print("缺少依赖：请先安装 bleak、openpyxl、pyserial，例如执行: pip install bleak openpyxl pyserial")
    except Exception as exc:
        print(f"扫描或连接失败：{exc}")


if __name__ == "__main__":
    main()
